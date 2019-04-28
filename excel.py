import openpyxl
import os
import re
import win32com.client as win32
os.chdir("C:\\Users\\Radek\\Documents\\FolderDoTestow\\Jacobs")
wb = openpyxl.load_workbook("LO_CH2M-UK.xlsx")
ws = wb["Sheet2"]

mailRegex = re.compile(r''' 
#name.surname@jacobs.com
[a-zA-Z0-9_\.]+ #name.surname
@               #@
[a-zA-Z0-9_\.]+  #jacobs.com
''', re.VERBOSE)
personRegex = re.compile(r'''
#surname, Name/(moze ale nie musi)(3 litery)

([a-zA-Z ]+ #surname
,)
([a-zA-Z ]+) #name
(/
[a-zA-Z ])? #kod biura
''', re.VERBOSE)
invoiceRegex = re.compile(r'''
([a-zA-Z0-9_\-/\. ]+) #invoice number and gibbersh 
(\d+\.\d\d) #invoice amount
''', re.VERBOSE)


people = []
for row in ws.iter_rows(min_row=1, max_col=2):
    for cell in row:
        if personRegex.search(str(cell.value)):
            people.append({"name": cell.value, "invoices": [], "email": []})
        if invoiceRegex.search(str(cell.value)):
            (people[-1])["invoices"].append(cell.value)
        if mailRegex.search(str(cell.value)):
            (people[-1])["email"].append(cell.value)
newLine = "\n"
# print(people[1]["invoices"][1])
for person in people:
    # i = 0??
    # print(person["name"])
    invoices = [] #to zdecydowanie mozna lepiej przeiterowac
    for invoice in person["invoices"]:
        invoices.append(invoice +"\n")
    #     print(invoice)
    # print(person["email"][0])    


    if len(invoices) == 1:
        message = (
            f"""Dear {personRegex.search(str(person['name'])).group(2).strip()}\n\n"""
            f"""I am contacting you because you have {len(invoices)} invoice waiting for your action in Liquid Office.\nThis invoice waits to be approved already for more than 90 days. The oldest invoice in your inbox is listed below.\n\n"""
            f"""Please login to the Liquid Office system and take the appropriate action to clear this item as soon as possible. If you are not able to do that, or believe that the information below is incorrect, please let us know about that.\nOver 90 days:\n\n"""
            f"""{''.join(invoices)}"""
            f"""\n\nIf you are on the Jacobs intranet the URL is http://liquidoffice.jacobs.com\n"""
            f"""If you are not on the Jacobs intranet the URL is http://connect.jacobs.com and use the Liquid Office link there.\n\n"""
            f"""If you have any technical issue preventing you from taking action on this invoice please contact jidsliquidoffice@jacobs.com.\n\n""" 
            f"""Regards,\n\nRadoslaw Gasior\nJacobs\nAccounting Professional | Accounts Payable\nradoslaw.gasior@jacobs.com""")   

    else:
        message = (
            f"""Dear {personRegex.search(str(person['name'])).group(2).strip()}\n\n"""
            f"""I am contacting you because you have {len(invoices)} invoices waiting for your action in Liquid Office.\nThose invoices wait to be approved already for more than 90 days. The oldest invoices in your inbox are listed below.\n\n"""
            f"""Please login to the Liquid Office system and take the appropriate action to clear those items as soon as possible. If you are not able to do that, or believe that the information below is incorrect, please let us know about that.\nOver 90 days:\n\n"""
            f"""{''.join(invoices)}"""
            f"""\n\nIf you are on the Jacobs intranet the URL is http://liquidoffice.jacobs.com\n"""
            f"""If you are not on the Jacobs intranet the URL is http://connect.jacobs.com and use the Liquid Office link there.\n\n"""
            f"""If you have any technical issue preventing you from taking action on those invoices please contact jidsliquidoffice@jacobs.com.\n\n""" 
            f"""Regards,\n\nRadoslaw Gasior\nJacobs\nAccounting Professional | Accounts Payable\nradoslaw.gasior@jacobs.com""")


    print(message)
    # outlook = win32.Dispatch("outlook.application")
    # mail = outlook.CreateItem(0)
    # mail.To = "radoslaw.gasior@jacobs.com"
    # mail.Subject = "ACTION REQUIRED: Outstanding invoices in Liquid Office"
    # mail.Body = message
    # mail.Send()

    # print(f"Dear {personRegex.search(str(person['name'])).group(2).strip()}{newLine}{newLine}You have {' '.join(invoices)}{newLine}{newLine}")


input()

#jezeli A1 spelnia kryteria to dodaj A1 jako dictionary key, nastepne wartosci
#w kolumnie A jako wartosci tego klucza (jako lista?) az kolejna wartosc nie spelni personRegex
#wtedy sprawdz czy C1 jest mailem ktory spelnia warunki i jezeli tak to dodaj
#go do listy.
#Maile beda przygotowywane jako Dear {wyciagniete imie z regex}, masz x faktur
#w inboxie, zajmij sie nimi, lista faktur powyzej 90 dni i lista faktur do 90 dni

#if jezeli kolejka to osobna lista!!!!!!!!!!!


#for num in range (1,131): # liczba wierszy
#



#pusta liste people w forze if z regexem od osoby, jezeli spelni warunek ze to osoba
#{name: czlowiek, faktury: [faktura1, faktura2], mail: mail@mail.com}
# jezeli znajdujesz goscia to tworzysz slownik, jezeli znajdujesz fakture to dodaj listy

        # if personRegex.search(str(cell.value)):
            

    # people.append([(personRegex.search(str(cell.value)).group()) for cell in row if personRegex.search(str(cell.value))])
    # people.append([(personRegex.search(str(cell.value)).group()) else (invoiceRegex.search(str(cell.value)).group()) for cell in row])?????????


        # if personRegex.search(str(cell.value)):
        #     print(cell.value)
            # print(personRegex.search(str(cell.value)).group(2).strip()) #zwraca imie danej osoby

#wyciaganie faktur
# for row in ws.iter_rows(min_row=1, min_col=1, max_col=1, max_row=131):
#     for cell in row:
#         if invoiceRegex.search(str(cell.value)):
#             # invoices.append(cell.value)
#             print(cell.value)   #zwraca fakture
#         # else:
#         #     print(cell.value) #zwraca człowieka lub kolejke
# #wyciaganie maili
# # mails = []
# for row in ws.iter_rows(min_row=1, max_col=3, max_row=131):
#     for cell in row:
#         if mailRegex.search(str(cell.value)):
#             print(cell.value)
# #             mails.append(mailRegex.search(str(cell.value)).group())
# print(mails)

#
