import openpyxl
import os
import re
import win32com.client as win32
os.chdir("C:\\Users\\Radek\\Documents\\FolderDoTestow\\Jacobs")
wb = openpyxl.load_workbook("LO_CH2M-UK.xlsx")
ws90 = wb["Sheet2"]
ws16 = wb["Sheet5"]

mailRegex = re.compile(r''' 
#name.surname@jacobs.com
[a-zA-Z0-9_\.]+ #name.surname
@               #@
[a-zA-Z0-9_\.]+  #jacobs.com
''', re.VERBOSE)
personRegex = re.compile(r'''
#surname, Name/(might contain, but not necessarily)(/3 letters)

([a-zA-Z ]+ #surname
,)
([a-zA-Z ]+) #name
([\/a-zA-Z]){3}? #kod biura
''', re.VERBOSE)

invoiceRegex = re.compile(r'''
([a-zA-Z0-9_\-\/\. ]+) #invoice number and gibbersh 
(\d+\.\d\d) #invoice amount
''', re.VERBOSE)

people = []
for row in ws90.iter_rows(min_row=1, max_col=2):
    for cell in row:
        if personRegex.search(str(cell.value)):
            people.append({"name": cell.value, "invoices90": [], "invoices16": [], "email": ""})
        if invoiceRegex.search(str(cell.value)):
            (people[-1])["invoices90"].append(cell.value)
        if mailRegex.search(str(cell.value)):
            (people[-1])["email"] = cell.value

for row in ws16.iter_rows(min_row=1, max_col=2):
    for cell in row:
        if personRegex.search(str(cell.value)):
            personExist = False
            for index, person in enumerate(people):
                if person["name"] == cell.value:
                    personExist = True
                    personID = index
                    break
            if personExist == False:
                people.append({"name": cell.value, "invoices90": [], "invoices16": [], "email": ""})
                personID = len(people) - 1
                print(people[personID])
        if invoiceRegex.search(str(cell.value)):
            people[personID]["invoices16"].append(cell.value)
            print(people[personID])
        if mailRegex.search(str(cell.value)):
            people[personID]["email"] = cell.value
            print(people[personID])

newLine = "\n"

for person in people: #po pipulach iterowac
    invoices90 = [] #to zdecydowanie mozna lepiej przeiterowac
    for invoice in person["invoices90"]:
        invoices90.append(invoice +"\n")
    invoices16 = []
    for invoice in person["invoices16"]:
        invoices16.append(invoice +"\n")

    if len(invoices90) > 0 and len(invoices16) > 0: #more then one old and more then one younger
        message = (
                f"""Dear {personRegex.search(str(person['name'])).group(2).strip()}\n\n"""
                f"""I am contacting you because you have {(len(invoices90) + len(invoices16))} invoices waiting for your action in Liquid Office.\nThose invoices wait to be approved already for more than 90 days. The oldest invoices in your inbox is listed below.\n\n"""
                f"""Please login to the Liquid Office system and take the appropriate action to clear those items as soon as possible. If you are not able to do that, or believe that the information below is incorrect, please let us know about that.\nOver 90 days:\n\n"""
                f"""{''.join(invoices90)}"""
                f"""{person["email"]}"""
                f"""\nOver 16 days:\n\n"""
                f"""{''.join(invoices16)}"""
                f"""\n\nIf you are on the Jacobs intranet the URL is http://liquidoffice.jacobs.com\n"""
                f"""If you are not on the Jacobs intranet the URL is http://connect.jacobs.com and use the Liquid Office link there.\n\n"""
                f"""If you have any technical issue preventing you from taking action on this invoice please contact jidsliquidoffice@jacobs.com.\n\n""" 
                f"""Regards,\n\nRadoslaw Gasior\nJacobs\nAccounting Professional | Accounts Payable\nradoslaw.gasior@jacobs.com""")   
    elif len(invoices90) == 1 and len(invoices16) == 0: #one old and no younger
        message = (
                f"""Dear {personRegex.search(str(person['name'])).group(2).strip()}\n\n"""
                f"""I am contacting you because you have 1 invoices waiting for your action in Liquid Office.\nThis invoice waits to be approved already for more than 90 days. The oldest invoice in your inbox is listed below.\n\n"""
                f"""Please login to the Liquid Office system and take the appropriate action to clear those items as soon as possible. If you are not able to do that, or believe that the information below is incorrect, please let us know about that.\nOver 90 days:\n\n"""
                f"""{''.join(invoices90)}"""
                f"""{person["email"]}"""
                f"""\n\nIf you are on the Jacobs intranet the URL is http://liquidoffice.jacobs.com\n"""
                f"""If you are not on the Jacobs intranet the URL is http://connect.jacobs.com and use the Liquid Office link there.\n\n"""
                f"""If you have any technical issue preventing you from taking action on this invoice please contact jidsliquidoffice@jacobs.com.\n\n""" 
                f"""Regards,\n\nRadoslaw Gasior\nJacobs\nAccounting Professional | Accounts Payable\nradoslaw.gasior@jacobs.com""")   
    elif len(invoices90) == 0 and len(invoices16) == 1: #no old and one younger
        message = (
                f"""Dear {personRegex.search(str(person['name'])).group(2).strip()}\n\n"""
                f"""I am contacting you because you have 1 invoices waiting for your action in Liquid Office.\nThis invoice waits to be approved already for more than 16 days. The oldest invoice in your inbox is listed below.\n\n"""
                f"""Please login to the Liquid Office system and take the appropriate action to clear this item as soon as possible. If you are not able to do that, or believe that the information below is incorrect, please let us know about that.\nOver 16 days:\n\n"""
                f"""{''.join(invoices16)}"""
                f"""{person["email"]}"""
                f"""\n\nIf you are on the Jacobs intranet the URL is http://liquidoffice.jacobs.com\n"""
                f"""If you are not on the Jacobs intranet the URL is http://connect.jacobs.com and use the Liquid Office link there.\n\n"""
                f"""If you have any technical issue preventing you from taking action on this invoice please contact jidsliquidoffice@jacobs.com.\n\n""" 
                f"""Regards,\n\nRadoslaw Gasior\nJacobs\nAccounting Professional | Accounts Payable\nradoslaw.gasior@jacobs.com""")   
    elif len(invoices90) > 1 and len(invoices16) == 0: #more then one old and no younger
        message = (
                f"""Dear {personRegex.search(str(person['name'])).group(2).strip()}\n\n"""
                f"""I am contacting you because you have {len(invoices90)} invoices waiting for your action in Liquid Office.\nThose invoices wait to be approved already for more than 90 days. The oldest invoices in your inbox is listed below.\n\n"""
                f"""Please login to the Liquid Office system and take the appropriate action to clear those items as soon as possible. If you are not able to do that, or believe that the information below is incorrect, please let us know about that.\nOver 90 days:\n\n"""
                f"""{''.join(invoices90)}"""
                f"""{person["email"]}"""
                f"""\n\nIf you are on the Jacobs intranet the URL is http://liquidoffice.jacobs.com\n"""
                f"""If you are not on the Jacobs intranet the URL is http://connect.jacobs.com and use the Liquid Office link there.\n\n"""
                f"""If you have any technical issue preventing you from taking action on this invoice please contact jidsliquidoffice@jacobs.com.\n\n""" 
                f"""Regards,\n\nRadoslaw Gasior\nJacobs\nAccounting Professional | Accounts Payable\nradoslaw.gasior@jacobs.com""")   
    elif len(invoices90) == 0 and len(invoices16) > 1: #more then one young and no older
        message = (
                f"""Dear {personRegex.search(str(person['name'])).group(2).strip()}\n\n"""
                f"""I am contacting you because you have {len(invoices16)} invoices waiting for your action in Liquid Office.\nThose invoices wait to be approved already for more than 16 days. The oldest invoices in your inbox is listed below.\n\n"""
                f"""Please login to the Liquid Office system and take the appropriate action to clear those items as soon as possible. If you are not able to do that, or believe that the information below is incorrect, please let us know about that.\nOver 16 days:\n\n"""
                f"""{''.join(invoices16)}"""
                f"""{person["email"]}"""
                f"""\n\nIf you are on the Jacobs intranet the URL is http://liquidoffice.jacobs.com\n"""
                f"""If you are not on the Jacobs intranet the URL is http://connect.jacobs.com and use the Liquid Office link there.\n\n"""
                f"""If you have any technical issue preventing you from taking action on this invoice please contact jidsliquidoffice@jacobs.com.\n\n""" 
                f"""Regards,\n\nRadoslaw Gasior\nJacobs\nAccounting Professional | Accounts Payable\nradoslaw.gasior@jacobs.com""")   
    print(message)

    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = "radoslaw.gasior@jacobs.com"
    mail.Subject = "ACTION REQUIRED: Outstanding invoices in Liquid Office"
    mail.Body = message
    mail.Send()

input()

